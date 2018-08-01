"""
Created on Thu Aug 10 15:44:36 2017

@author: agnisha.singh
"""
from Tkinter import *
import selenium
import Tkinter as tk
import win32gui
from unidecode import unidecode
import random
import time
import unicodedata
import subprocess
import string
from openpyxl.styles import Style, Color, Font, PatternFill
import openpyxl.styles.colors as colors
import traceback
import tkFileDialog as filedialog
from fuzzywuzzy import fuzz
import sys
from selenium.webdriver.common.keys import Keys
from os.path import expanduser
import os
import xlsxwriter
import openpyxl
from xlrd import *
import win32com.client  
import tkMessageBox
import os
from selenium import webdriver
import re
home = expanduser("~")

def get_reg_no_diff_list(registry_number_list,prefile_registry_no_to_be_written):
    reg_no_match_list=[]
    
    if len(registry_number_list)==len(prefile_registry_no_to_be_written):
        for i in range(len(registry_number_list)):
          if prefile_registry_no_to_be_written[i] is not "":  
            if registry_number_list[i]==prefile_registry_no_to_be_written[i]:
                reg_no_match_list.append("MATCH")
            else:
                reg_no_match_list.append("NO MATCH")
          else:
              continue
    return reg_no_match_list

# methods to get the differences list (MATCH/NO MATCH) between fetched data and pre-file data
def get_differences_list(fetched_list_to_be_sent_for_differences,prefile_list_to_sent_for_differences):        
    official_entity_name_match_list=[]
    address_match_list=[]
    legal_form_match_list=[]
    registry_no=[]
    
#    make a unique list of registry numbers from fetched list removing duplicates and maintaing the sequence
    for item in fetched_list_to_be_sent_for_differences[0]:
        if item not in registry_no:
            registry_no.append(item)

    previous_names_match_list=[]
    match_list_to_be_sent=[official_entity_name_match_list,address_match_list,legal_form_match_list,previous_names_match_list]
    check_list=[]
#        make a check list of checked fields in the tool
    if CheckVar2.get() is 1:
        check_list.append(0)
    if CheckVar3.get() is 1: 
        check_list.append(1)
    if CheckVar4.get() is 1:
        check_list.append(2)
    
#       for each checked item get the match list
    for j in range(0,3):
      if j in check_list:              
        for i in range(len(prefile_list_to_sent_for_differences[0])):         
          
           if re.sub(r"\W", "",str(fetched_list_to_be_sent_for_differences[j+1][i]).lower())==re.sub(r"\W", "",str(prefile_list_to_sent_for_differences[j+1][i]).lower()):
              match_list_to_be_sent[j].append("MATCH")                      
           else:
              match_list_to_be_sent[j].append("NO MATCH")

    if  CheckVar6.get() is 1:  
         for i in range(len(prefile_list_to_sent_for_differences[0])):
            
             if 'No previous name information has been recorded over the last 20 years.' in fetched_list_to_be_sent_for_differences[4][i] and str(prefile_list_to_sent_for_differences[4][i].lstrip().rstrip()) == "":
                  match_list_to_be_sent[3].append("MATCH")
             else:     
                 if re.sub(r"\W", "",str(fetched_list_to_be_sent_for_differences[4][i]).lower())==re.sub(r"\W", "",str(prefile_list_to_sent_for_differences[4][i] ).lower()):
                   match_list_to_be_sent[3].append("MATCH")
                 else:
                   match_list_to_be_sent[3].append("NO MATCH")  
 
    
    
    
    return match_list_to_be_sent



#method to read data from pre-file    
def read_pre_file(file_path):
    input_workbook=open_workbook(file_path)
    sheet = input_workbook.sheet_by_index(0)
    column_names=['BusinessRegistryCountry','OfficialBusinessRegistryReference','OfficialEntityName','LegalFormationAddressLine1', 'LegalFormationAddressLine2', 'LegalFormationAddressTown/City', 'LegalFormationAddressCounty/State','LegalFormationAddressPostCode', 'LegalForm', 'Previous Entity Name']
    col_no=sheet.ncols
    required_list=[]        
    for column_name in column_names:
      for i in range(1,col_no) :       
        if (sheet.cell(0,i).value)== column_name:
            required_list.append(i)           
            break
 
    prefile_country_list=[]
    prefile_registry_no_list=[]
    prefile_entity_name_list=[]
    prefile_address_list=[]
    prefile_legal_form_list=[]
    prefile_prev_name_list=[]
   
   
    for i in range(1,sheet.nrows):
        prefile_country_list.append(sheet.cell(i,required_list[0]).value)
        prefile_registry_no_list.append(sheet.cell(i,required_list[1]).value)
        prefile_entity_name_list.append(sheet.cell(i,required_list[2]).value)
        prefile_address_list.append(sheet.cell(i,required_list[3]).value+' '+sheet.cell(i,required_list[4]).value+' '+sheet.cell(i,required_list[5]).value+' '+sheet.cell(i,required_list[6]).value+' '+sheet.cell(i,required_list[7]).value)
        prefile_legal_form_list.append(sheet.cell(i,required_list[8]).value) 
        prefile_prev_name_list.append(sheet.cell(i,required_list[9]).value)
    prefile_data_list=[prefile_country_list,prefile_registry_no_list,prefile_entity_name_list,prefile_address_list,prefile_legal_form_list,prefile_prev_name_list]    
    return prefile_data_list    

#method to convert column number to column title
def convertToTitle(num):
    title = ''
    alist = string.uppercase
    while num:
        mod = (num-1) % 26
        num = int((num - mod) / 26)  
        title += alist[mod]
    return title[::-1]   

#method to write fetched data, prefile data and the differnces data in the result Excel file       
def write_data_to_excel(list_to_send_write_function, count,check_box_states):
    was_File_existing=False
    
#    make background fills
    blueFill = PatternFill(start_color='5AB7E8',
                   end_color='5AB7E8',
                   fill_type='solid')
    purpleFill= PatternFill(start_color='9D2C7D',
                   end_color='9D2C7D',
                   fill_type='solid') 
    greyFill= PatternFill(start_color='939598',
                   end_color='939598',
                   fill_type='solid')  
    evs_red_fill= PatternFill(start_color='EE2653',
                   end_color='EE2653',
                   fill_type='solid') 
    greenFill= PatternFill(start_color='00B050',
                   end_color='00B050',
                   fill_type='solid')  
    lightgreyFill= PatternFill(start_color='F1CFE8',
                   end_color='F1CFE8',
                   fill_type='solid')      
    try:
       try:
           os.remove("{}/Desktop/Entity Data Validation-Output.xlsx".format(home))
           was_File_existing= True
       except: 
           pass 
       
       data_file=openpyxl.Workbook()
       sheet=data_file.worksheets[0]
       sheet.title="United Kingdom"
#       data_worksheet=data_file.worksheets[0]
       country="United Kingdom"  
#        headers for fetched
       headers_format_list=['REGISTRY NUMBER', 'OFFICIAL ENTITY NAME', 'LEGAL ADDRESS', 'LEGAL FORM', 'STATUS','DISSOLVE DATE', \
       'PREVIOUS NAMES','DATES OF NAME CHANGE']
       col=1
       for header in headers_format_list:
            sheet.cell(row=1, column=col).value =header             
            sheet.cell(row=1, column=col).font =sheet.cell(row=1, column=col).font.copy(bold=True)            
            sheet.cell(row=1, column=col).style = Style(font=Font(color=Color(colors.WHITE)))
            sheet.cell(row=1, column=col).fill=blueFill
            col=col+1
#        headers for pre
       headers_format_list=['REGISTRY NUMBER', 'OFFICIAL ENTITY NAME', 'LEGAL ADDRESS', 'LEGAL FORM','PREVIOUS NAMES']     
       for header in headers_format_list:
            sheet.cell(row=1, column=col).value =header             
            sheet.cell(row=1, column=col).font =sheet.cell(row=1, column=col).font.copy(bold=True)            
            sheet.cell(row=1, column=col).style = Style(font=Font(color=Color(colors.WHITE)))
            sheet.cell(row=1, column=col).fill=purpleFill
            col=col+1
            
#        headers for matched
     
       headers=['REGISTRY NUMBER','OFFICIAL ENTITY NAME', 'LEGAL ADDRESS', 'LEGAL FORM','PREVIOUS NAMES']
       for i in range(len(headers)):
            sheet.cell(row=1, column=col).value =headers[i]
            sheet.cell(row=1, column=col).font = sheet.cell(row=1, column=col).font.copy(bold=True)            
            sheet.cell(row=1, column=col).style = Style(font=Font(color=Color(colors.WHITE)))
            sheet.cell(row=1, column=col).fill=greyFill
            col=col+1 
            
       #fetched data lists            
       registry_number_list=  list_to_send_write_function[0]         
       official_entity_name_list=list_to_send_write_function[1]
       legal_address_list=list_to_send_write_function[2]
       legal_form_list=list_to_send_write_function[3]
       status_list=list_to_send_write_function[4]
       dissolve_date_list=list_to_send_write_function[5]
       previous_names_list=list_to_send_write_function[6]
       date_of_change_list=list_to_send_write_function[7]  
       statusVar.set("Reading Excel data...")
       top.update()
       #data lists from pre-file            
       prefile_data= read_pre_file (file_path)

       prefile_country_list=prefile_data[0]
       prefile_registry_no_list_with_spaces=prefile_data[1]
       prefile_registry_no_list=[]
       for no in prefile_registry_no_list_with_spaces:
            no=no.lstrip().rstrip()
            prefile_registry_no_list.append(no)
       prefile_entity_name_list=prefile_data[2]
       prefile_address_list=prefile_data[3]       
       prefile_legal_form_list=prefile_data[4]
       prefile_prev_name_list=prefile_data[5]
      
       prefile_registry_no_to_be_written=[]
       prefile_entity_name_to_be_written=[]
       prefile_address_to_be_written=[]
       prefile_legal_form_to_be_written=[]
       prefile_prev_name_to_be_written=[]

       for i in range(len( prefile_country_list)):
            if str(prefile_country_list[i]).lower()== country.lower():                
                prefile_registry_no_to_be_written.append(prefile_registry_no_list[i])
                prefile_entity_name_to_be_written.append(prefile_entity_name_list[i])
                prefile_address_to_be_written.append(prefile_address_list[i])
                prefile_legal_form_to_be_written.append(prefile_legal_form_list[i])
                prefile_prev_name_to_be_written.append(prefile_prev_name_list[i])
        
       if len(prefile_registry_no_list)!=count:
            prefile_registry_no_to_be_written=prefile_registry_no_to_be_written[0:count]
            prefile_entity_name_to_be_written=prefile_entity_name_to_be_written[0:count]
            prefile_address_to_be_written=prefile_address_to_be_written[0:count]
            prefile_legal_form_to_be_written=prefile_legal_form_to_be_written[0:count]
            prefile_prev_name_to_be_written=prefile_prev_name_to_be_written[0:count]
            
       prefile_list_to_sent_for_differences=[prefile_registry_no_to_be_written,prefile_entity_name_to_be_written,prefile_address_to_be_written,prefile_legal_form_to_be_written,prefile_prev_name_to_be_written]
       fetched_list_to_be_sent_for_differences=[registry_number_list,official_entity_name_list,legal_address_list,legal_form_list,previous_names_list]                       
        
       differences_list=get_differences_list(fetched_list_to_be_sent_for_differences,prefile_list_to_sent_for_differences)
        
       registry_numbers_differences_list=get_reg_no_diff_list(registry_number_list,prefile_registry_no_to_be_written) 
       statusVar.set("Exporting data to Excel...")
       top.update()
 #        write fetched data

       global_row=sheet.max_row+1
       row=global_row
       for i in range(len(registry_number_list)):
            
            if check_box_states[0] is True:
                sheet.cell(row=row, column=1).value = registry_number_list[i]
                sheet.column_dimensions[convertToTitle(1)].width = 18
                
            if check_box_states[1] is True: 
                sheet.cell(row=row, column=2).value =official_entity_name_list[i]
                sheet.column_dimensions[convertToTitle(2)].width = 22
                
            if check_box_states[2] is True: 
                sheet.cell(row=row, column=3).value =legal_address_list[i]
                sheet.column_dimensions[convertToTitle(3)].width = 15.5
                
            if check_box_states[3] is True:
                sheet.cell(row=row, column=4).value =legal_form_list[i]
                sheet.column_dimensions[convertToTitle(4)].width = 13.5
                
            if check_box_states[4] is True: 
                status,sep,deletion_date=status_list[i].partition(",")                
                sheet.cell(row=row, column=5).value =status
                sheet.column_dimensions[convertToTitle(5)].width = 10
                
            if check_box_states[4] is True:
                sheet.cell(row=row, column=6).value =dissolve_date_list[i]
                sheet.column_dimensions[convertToTitle(6)].width = 15
                
            if check_box_states[5] is True:
                sheet.cell(row=row, column=7).value =previous_names_list[i]
                sheet.column_dimensions[convertToTitle(7)].width = 17
                
            if check_box_states[6] is True:
                sheet.cell(row=row, column=8).value =date_of_change_list[i]
                sheet.column_dimensions[convertToTitle(8)].width = 23.5

            row +=1     

       row=global_row
       for i in range(len(prefile_registry_no_to_be_written)):           
            if check_box_states[0] is True:                
                col=9               
                sheet.cell(row=row, column=col).value = prefile_registry_no_to_be_written[i]
                sheet.column_dimensions[convertToTitle(col)].width = 18
                                
            if check_box_states[1] is True:                 
                col=10                
                sheet.cell(row=row, column=col).value =prefile_entity_name_to_be_written[i]
                sheet.column_dimensions[convertToTitle(col)].width = 22
                
            if check_box_states[2] is True:
                col=11               
                sheet.cell(row=row, column=col).value =prefile_address_to_be_written[i]
                sheet.column_dimensions[convertToTitle(col)].width = 16
                
            if check_box_states[3] is True:
                col=12
                sheet.cell(row=row, column=col).value =prefile_legal_form_to_be_written[i]
                sheet.column_dimensions[convertToTitle(col)].width = 15
                
            if check_box_states[5] is True :             
                sheet.cell(row=row, column=13).value =prefile_prev_name_to_be_written[i]
                sheet.column_dimensions[convertToTitle(13)].width = 17
            row +=1    
       
            
# write match data        
       row=global_row
       for i in range(len(registry_numbers_differences_list)):            
            if check_box_states[0] is True:                    
                col=14                 
                sheet.cell(row=row, column=col).value =  registry_numbers_differences_list[i] 
                sheet.column_dimensions[convertToTitle(col)].width = 18
                if str(registry_numbers_differences_list[i])=='MATCH':
                    sheet.cell(row=row, column=col).style = Style(font=Font(color=Color(colors.WHITE)))
                    sheet.cell(row=row, column=col).fill=greenFill
                elif str(registry_numbers_differences_list[i])=='NO MATCH':
                    sheet.cell(row=row, column=col).style = Style(font=Font(color=Color(colors.WHITE)))
                    sheet.cell(row=row, column=col).fill=evs_red_fill
            row+=1
       row=global_row    

       for i in range(len(prefile_registry_no_to_be_written)):
                if check_box_states[1] is True:                    
                    col=15                    
                    sheet.cell(row=row, column=col).value =differences_list[0][i]
                    sheet.column_dimensions[convertToTitle(col)].width = 22
                    if str(differences_list[0][i])=='MATCH':
                        sheet.cell(row=row, column=col).style = Style(font=Font(color=Color(colors.WHITE)))
                        sheet.cell(row=row, column=col).fill=greenFill
                    elif str(differences_list[0][i])=='NO MATCH':
                        sheet.cell(row=row, column=col).style = Style(font=Font(color=Color(colors.WHITE)))
                        sheet.cell(row=row, column=col).fill=evs_red_fill
                    
                if check_box_states[2] is True: 
                    col=16
                    sheet.cell(row=row, column=col).value =differences_list[1][i]
                    sheet.column_dimensions[convertToTitle(col)].width = 15.5
                    if str(differences_list[1][i])=='MATCH':
                        sheet.cell(row=row, column=col).style = Style(font=Font(color=Color(colors.WHITE)))
                        sheet.cell(row=row, column=col).fill=greenFill
                    elif str(differences_list[1][i])=='NO MATCH':
                        sheet.cell(row=row, column=col).style = Style(font=Font(color=Color(colors.WHITE)))
                        sheet.cell(row=row, column=col).fill=evs_red_fill
            
                if check_box_states[3] is True:
                    
                    col=17
                    
                    sheet.cell(row=row, column=col).value =differences_list[2][i]
                    sheet.column_dimensions[convertToTitle(col)].width = 13.5
                    if str(differences_list[2][i])=='MATCH':
                        sheet.cell(row=row, column=col).style = Style(font=Font(color=Color(colors.WHITE)))
                        sheet.cell(row=row, column=col).fill=greenFill
                    elif str(differences_list[2][i])=='NO MATCH':
                        sheet.cell(row=row, column=col).style = Style(font=Font(color=Color(colors.WHITE)))
                        sheet.cell(row=row, column=col).fill=evs_red_fill

                    
                if check_box_states[5] is True :
                    sheet.cell(row=row, column=18).value =differences_list[3][i]
                    sheet.column_dimensions[convertToTitle(18)].width = 22
                    if str(differences_list[3][i])=='MATCH':
                        sheet.cell(row=row, column=18).style = Style(font=Font(color=Color(colors.WHITE)))
                        sheet.cell(row=row, column=18).fill=greenFill
                    elif str(differences_list[3][i])=='NO MATCH':
                        sheet.cell(row=row, column=18).style = Style(font=Font(color=Color(colors.WHITE)))
                        sheet.cell(row=row, column=18).fill=evs_red_fill                                        
                row+=1              
       data_file.save("{}/Desktop/Entity Data Validation-Output.xlsx".format(home))
       return was_File_existing
     
    except IOError:
        tkMessageBox.showerror("Output File Open","Your Output File was open. Close it and run tool again!")        
        sys.exitfunc()         

#method to crawl UK
def crawl_uk(browser):
 try:       
    #define lists
    
    check_box_states=[False,False,False,False,False,False,False]
    if CheckVar4.get() is 1:
        check_box_states[3]=True
    if CheckVar1.get() is 1:
       check_box_states[0]=True 
    if CheckVar2.get() is 1: 
       check_box_states[1]=True 
    if CheckVar3.get() is 1:
       check_box_states[2]=True
    if CheckVar5.get() is 1: 
       check_box_states[4]=True 
    if CheckVar6.get() is 1: 
       check_box_states[5]=True 
    if CheckVar7.get() is 1: 
       check_box_states[6]=True    
    registry_number_list=[] 
    official_entity_name_list=[]  
    legal_address_list=[]
    legal_form_list=[]
    status_list=[]
    dissolve_date_list=[]
    previous_names_list=[]
    date_of_change_list=[]  
    registry_number_input_list=registry_no_list  
               
    if len(registry_number_input_list) is 0:
        return
#    statusVar.set("Extracting data")
#    top.update()
    #crawl for each input
    cnt=1  
   
    for registry_number_input in registry_number_input_list: 

      a= registry_number_input
      try:  
        statusVar.set("Extracting data("+str(cnt)+"/"+str(len(registry_number_input_list))+") :"+str(a.encode("utf8")))
        
        top.update()
        
        registry_number_input=registry_number_input.rstrip().strip()
        time.sleep(3)
        browser.get('http://wck2.companieshouse.gov.uk/')  
        searchTermElement=browser.find_element_by_id("cnumb")
        searchTermElement.clear()
        searchTermElement.send_keys(registry_number_input)
        searchButton=browser.find_element_by_xpath("//div[@class='searchButton']/input[@name='cosearch']")
        browser.execute_script("arguments[0].click();", searchButton)
        time.sleep(3)
        try:
            error_element=browser.find_element_by_class_name("error")
            if error_element is not None:
               registry_number_list.append(a)
               official_entity_name_list.append("Invalid Registry Number")
               legal_address_list.append("Invalid Registry Number")
               legal_form_list.append("Invalid Registry Number")
               status_list.append("Invalid Registry Number")
               dissolve_date_list.append("Invalid Registry Number")
               previous_names_list.append("Invalid Registry Number")
               date_of_change_list.append("Invalid Registry Number")
#               browser.quit()
               continue
        except:  
            officialEntityNameElement=browser.find_element_by_xpath("//td[@class='padding36']/strong")
            officialEntityName=officialEntityNameElement.text 
            if CheckVar2.get() is 1: 
                check_box_states[1]=True
                if officialEntityName is not None:
                    official_entity_name_list.append(str(officialEntityName))
                else:
                    official_entity_name_list.append("Unavailable")
                                   
            legalAddressElements=browser.find_elements_by_xpath("//td[@class='padding36']")
            for legalAddressElement in legalAddressElements:
                if legalAddressElement.text is not None:
                    legalAddress=legalAddressElement.text
                    legalAddress=str(legalAddress)
                    head,sep,tail= legalAddress.partition(':')
                    head,sep,tail=tail.partition('Company')
                    legalAddress=str(head.strip())
                    if legalAddress.startswith(str(officialEntityName)):
                        legalAddress=legalAddress[len(officialEntityName):]
                        legalAddress=legalAddress.strip()
                    head,sep,tail=tail.partition('.')
                    registryNumber=tail.strip()               
    
            if CheckVar1.get() is 1:
                check_box_states[0]=True            
                if registryNumber is not None:
                    registry_number_list.append(str(registryNumber))
                else:
                    registry_number_list.append("Unavailable")            
            if CheckVar3.get() is 1: 
                check_box_states[2]=True
                if legalAddress is not None:
                    legal_address_list.append(str(legalAddress))
                else:
                    legal_address_list.append("Unavailable")
            
            
            tableElements=browser.find_elements_by_xpath("//td[@class='yellowCreamTable']")          
            legal_form_row=str(tableElements[1].text)
            head,sep,tail=legal_form_row.partition(':')
            head,sep,tail=tail.partition('Nature')  
            legal_form=head.strip()
    
            
            if CheckVar4.get() is 1:
                check_box_states[3]=True
                if legal_form is not None:
                    legal_form_list.append(str(legal_form))
                else:
                    legal_form_list.append("Unavailable")
            
            
            status_row=str(tableElements[0].text)        
            head,sep,tail=status_row.partition(':')
            head,sep,tail=tail.partition('Date')
            status=head.strip()
            dissolve_date="-"
            if 'Dissolved' in status:
                head,sep,tail=status.partition(' ')
                status=head
                dissolve_date=tail           
    
                
            if CheckVar5.get() is 1: 
                check_box_states[4]=True
                if status is not None:
                    status_list.append(str(status))
                else:
                    status_list.append("Unavailable")
    
                if dissolve_date is not None:
                    dissolve_date_list.append(str(dissolve_date))
                else:
                    dissolve_date_list.append("Unavailable")
             
            
            previous_names_element=tableElements[4].text        
            head,sep,tail= previous_names_element.partition(':') 
            entire_text=tail
            prev_names=[]
            dates_of_change=[]
            final_dates=""
            final_prev_names=""
            if 'Date' in tail:
                head,sep,tail=tail.partition('Name')
                remaining_text=tail
                match_dates=re.findall(r'\d{2}/\d{2}/\d{4}', tail)
                dates_of_change=match_dates
                for i in range(len(dates_of_change)-1):
                    
                    head,sep,tail=remaining_text.partition(str(dates_of_change[i]))
                    head,sep,tail=tail.partition(str(dates_of_change[i+1]))
                    prev_name=head
                    prev_names.append(prev_name)
                head,sep,tail= entire_text.partition(str(dates_of_change[len(dates_of_change)-1])) 
                prev_names.append(tail)
                date_no=0
                for date in dates_of_change:
                    if date_no==0:
                        final_dates=str(date)
                    else:    
                        final_dates=final_dates+","+str(date)
                    date_no+=1  
                name_no=0    
                for name in prev_names:
                    if name_no==0:
                        final_prev_names=str(name)
                    else:    
                        final_prev_names=final_prev_names+","+str(name)
                    name_no+=1     

            else:
                final_prev_names=str(entire_text.strip())
                final_dates='-'
    
            if CheckVar6.get() is 1:
                check_box_states[5]=True
                previous_names_list.append(final_prev_names)
       
            if CheckVar7.get() is 1:
                check_box_states[6]=True
                date_of_change_list.append(final_dates)
               
        cnt+=1    
      except:
         print a 
         print cnt

         print traceback.print_exc()

    official_entity_name_list=[official_entity.decode('utf-8') for official_entity in official_entity_name_list] 
    official_entity_name_list=[official_entity.encode('ascii', 'ignore').decode('ascii') for official_entity in official_entity_name_list]
    list_to_send_write_function=[registry_number_list,official_entity_name_list,legal_address_list,legal_form_list,status_list,dissolve_date_list,previous_names_list,date_of_change_list]
    
    count=len(registry_number_input_list)
    print count
    if count !=len(list_to_send_write_function[0]):
        return "Entered exception"
    was_file_existing=write_data_to_excel(list_to_send_write_function,count,check_box_states)
    return was_file_existing
#end of crawl UK
 except:
     error_msg= "Entered exception"
     list_to_send_write_function=[registry_number_list,official_entity_name_list,legal_address_list,legal_form_list,status_list,dissolve_date_list,previous_names_list,date_of_change_list]    
 
     return [error_msg, list_to_send_write_function,check_box_states]
     
def uncheckFields_after_execution():
    upload_button_text_variable.set("Upload Input File") 
    C6.config(state=NORMAL)
    C7.config(state=NORMAL)
    C1.deselect()
    C2.deselect()
    C3.deselect()
    C4.deselect()
    C5.deselect()
    C6.deselect()
    C7.deselect()
    upload_button_text_variable.set("Upload Input File") 
    runTool_button_text_variable.set("Run")
    statusVar.set("Process Complete!")
#    root.destroy()
    
def viewOutput():
    xlApp = win32com.client.Dispatch("Excel.Application")
    xlApp.Visible = True
    try:
       xlApp.Workbooks.Open('{}/Desktop/Entity Data Validation-Output.xlsx'.format(home))
    except:  
        tkMessageBox.showerror("File already open", "Result file is already open or does not exist.")  
        
 
def uncheckFields():
   upload_button_text_variable.set("Upload Input File") 
   C5.config(state=NORMAL)
   C6.config(state=NORMAL)
   C7.config(state=NORMAL)
   if UK_Var.get() is not 1:
       C6.config(state=DISABLED)
       C7.config(state=DISABLED)
       
       
def runTool():    
  try:  
    upload_button_text=upload_button_text_variable.get()
   
    if not(str(upload_button_text) =='File Uploaded!'):
        tkMessageBox.showerror("No File Uploaded", "Please upload a file.")
        return  
    result=tkMessageBox.askyesno("Entity Data Validation-UK","Output file must be closed if any. Proceed?") 
    
    if result==True:    
        if CheckVar1.get() is 0 and CheckVar2.get() is 0 and CheckVar3.get() is 0 and CheckVar4.get() is 0 and CheckVar5.get() is 0 and CheckVar6.get() is 0 and CheckVar7.get() is 0:
                tkMessageBox.showerror("No field selected","Please select fields !")
                return
        if result is True:
            runTool_button_text_variable.set("Running...")
            top.update()        
        if getattr(sys,'frozen',False):
            exePath=os.path.dirname(os.path.realpath(sys.executable))
        elif __file__:
            exePath=os.path.dirname(os.path.abspath(__file__))
        dirPth= exePath+"\\chromedriver.exe"    
        browser=webdriver.Chrome(executable_path=dirPth)
        if len(registry_no_list) is not 0:    
                was_File_existing=crawl_uk(browser)
       
    #    in case of errors follow this method to write whatever is crawled        
        if type(was_File_existing) !=bool : 
            print "entered here"

            if was_File_existing[0]=="Entered exception":
                len_list=[]
                final_fetched_list=[]
                fetched_exception_list=was_File_existing[1]
                check_box_states=was_File_existing[2]
                              
                range_of_iteration=7 
                for i in range(range_of_iteration):
                  if check_box_states[i]==True:  
                    len_list.append(len(was_File_existing[1][i]))
                min_len=min(len_list)
                
                final_fetched_list=[fetched_exception_list[0][0:min_len],fetched_exception_list[1][0:min_len],fetched_exception_list[2][0:min_len],
                                  fetched_exception_list[3][0:min_len],fetched_exception_list[4][0:min_len],fetched_exception_list[5][0:min_len],
                                  fetched_exception_list[6][0:min_len],fetched_exception_list[7][0:min_len]] 
               
                count=min_len
                
                if count!=0:
                  write_data_to_excel(final_fetched_list, count,check_box_states)
                else:
                  traceback.print_exc()
                  tkMessageBox.showerror("Error", "The tool could not fetch any data! Please try again.")    
                  root.destroy()
                  sys.exit()
        browser.quit() 
    #        browser.close()
        if was_File_existing ==True:
                tkMessageBox.showinfo("Successful Execution", "Result file successfully updated on Desktop!")
        else:   

                tkMessageBox.showinfo("Successful Execution", "Result file successfully created on Desktop!")       
        uncheckFields_after_execution()        
  except:
    traceback.print_exc()
#    wait=input("123")      
        
        
def uploadFile() :
    try:        
        global registry_no_list
        global file_path  
        registry_no_input_list_UK=[] 
        file_path_input = filedialog.askopenfilename()               
        file_path=file_path_input
        input_workbook=open_workbook(file_path)
        sheet = input_workbook.sheet_by_index(0)
        for i in range(sheet.ncols):
            if sheet.cell(0,i).value=='OfficialBusinessRegistryReference':
                registry_no_col=i
                break
        for i in range(sheet.ncols):
            if sheet.cell(0,i).value== 'BusinessRegistryCountry':
                country_col=i
                break  
            
 #       verify if the selected country has any registry numbers in the pre file or not     
        if registry_no_col ==0: 
            tkMessageBox.showerror("No Registry Numbers","Input file does not contain registry numbers! Upload file again.")
            upload_button_text_variable.set("Upload Input File")
            return
        if country_col ==0: 
            tkMessageBox.showerror("No Country Names","Input file doesnot contain country names! Upload file again.")
            upload_button_text_variable.set("Upload Input File")
            return 
            
        end_of_registry_no_row=sheet.nrows
                
        for i in range(1,end_of_registry_no_row):
            if str(sheet.cell(i,country_col).value).lower() == "United Kingdom".lower():
                registry_no_input_list_UK.append(sheet.cell(i,registry_no_col).value)
        if len(registry_no_input_list_UK)==0:
            tkMessageBox.showerror("No Registry Numbers", "Input file does not contain any registry numbers for UK.Upload another file.")
            return
        else:
            registry_no_list=registry_no_input_list_UK
        if registry_no_input_list_UK is not None:
            statusVar.set("Number of records found : "+str(len(registry_no_input_list_UK)))
            upload_button_text_variable.set("File Uploaded!") 
        
    except:
        tkMessageBox.showerror("File Upload Failed", "Please upload a valid file")
        traceback.print_exc()
        
#=========================main interface=========================================        
if __name__ == "__main__":
    try:
        root = tk.Tk()
        fname="evs1.ico"
        if getattr(sys,'frozen',False):
            exePath=os.path.dirname(os.path.realpath(sys.executable))
        elif __file__:
            exePath=os.path.dirname(os.path.abspath(__file__))
            
        iPth=exePath+"\\"+fname
        
        
        root.iconbitmap(iPth)
        
        dataValidationframe = Frame(root,width=400,height=150,bd=5,relief=GROOVE)
        
        #give title to the parent window
        root.title('Entity Data Validation-UK')
        dataValidationframe.pack()
        
        #define fields
        fieldsVar = StringVar()
        fieldsLabel = Label( dataValidationframe, textvariable=fieldsVar )
        fieldsLabel.place(x=25,y=10)
        fieldsVar.set("Select Fields :")
        
        #checkboxes
        CheckVar1 = IntVar()
        CheckVar2 = IntVar()
        CheckVar3 = IntVar()
        CheckVar4 = IntVar()
        CheckVar5 = IntVar()
        CheckVar6 = IntVar()
        CheckVar7 = IntVar()
        C1 = Checkbutton(dataValidationframe, text = "Registry Number", variable = CheckVar1, onvalue = 1, offvalue = 0)
        C2 = Checkbutton(dataValidationframe, text = "Official Entity Name", variable = CheckVar2, onvalue = 1, offvalue = 0)
        C3 = Checkbutton(dataValidationframe, text = "Legal Address", variable = CheckVar3, onvalue = 1, offvalue = 0)
        C4 = Checkbutton(dataValidationframe, text = "Legal Form", variable = CheckVar4, onvalue = 1, offvalue = 0)
        C5 = Checkbutton(dataValidationframe, text = "Status", variable = CheckVar5, onvalue = 1, offvalue = 0)
        C6 = Checkbutton(dataValidationframe, text = "Previous Name", variable = CheckVar6, onvalue = 1, offvalue = 0)
        C7 = Checkbutton(dataValidationframe, text = "Date of Name Change", variable = CheckVar7, onvalue = 1, offvalue = 0)
        C1.place(x=45,y=30)
        C2.place(x=195,y=30)
        C3.place(x=45,y=50)
        C4.place(x=195,y=50)
        C5.place(x=45,y=70)
        C6.place(x=195,y=70)
        C7.place(x=45,y=90)
        
        top = Frame(root)
        statusFrame=Frame(top,width=400,height=30,bd=5,relief=GROOVE)
        statusFrame.pack()
        global statusVar
        statusVar = StringVar()
        statusLabel = Label( statusFrame, textvariable=statusVar)
        statusLabel.place(x=0,y=0)
        statusLabel.configure(foreground="blue")
        statusVar.set("Status :  No file selected.")
        bottom = Frame(root)
        top.pack(side=TOP)
        bottom.pack(side=BOTTOM, fill=BOTH, expand=True)
        outputToolButton = Button(root, text="View Output", width=10, height=1, command=viewOutput)
        runTool_button_text_variable=StringVar()
        runToolButton = Button(root, textvariable=runTool_button_text_variable, width=10, height=1,command=runTool)
        runTool_button_text_variable.set("Run Tool")
        upload_button_text_variable=StringVar()
        upload_button_text_variable.set("Upload Input File")
        uploadloadTemplateButton=Button(root, textvariable=upload_button_text_variable, width=25, height=1,command=uploadFile)
        outputToolButton.pack(in_=top, side=RIGHT,padx=10)
        runToolButton.pack(in_=top, side=RIGHT,padx=5, pady=5)
        uploadloadTemplateButton.pack(in_=top, side=RIGHT,padx=10)
        root.mainloop()
    except:
        traceback.print_exc()
        wait=input("Error occurred")

    